from openpyxl import Workbook
from openpyxl import load_workbook
from package.Infrastructure import FileIO

fileObj = FileIO.FileIO('OTCBuySell', '20180727')

wb = load_workbook(filename=fileObj.SaveAs)
wbList = wb.sheetnames
#從後面取4個sheet
overFive = len(wbList) >= 5 and wbList[-4:] or []
print('共有幾個sheet:')
print(len(overFive))