# -*- coding: utf-8 -*-
# 買賣超前指定資料筆數輸出產生excel sheet物件
import sys
import copy
import numpy as np
from package.Infrastructure import DateObj, FileIO, HistoryFind
from package.Infrastructure import CommonMethon as _cm
# 使用openpyxl内置的格式
from openpyxl.styles import numbers, Font, PatternFill, colors
# enum multiple values
from aenum import MultiValueEnum


class ColorEnum(MultiValueEnum):
    """
    顏色代碼aenum
    """
    # 外資、投信同步 買或賣超
    Yello = 0, PatternFill(fill_type='solid', fgColor=colors.YELLOW)
    # 外資、投信不同步 買或賣超
    Green = 1, PatternFill(fill_type='solid', fgColor=colors.GREEN)
    # 5日內累積買或賣超 2天
    Cnt2Days = 2, PatternFill(
        fill_type='solid', fgColor=colors.Color(rgb='95E1D3'))
    # 5日內累積買或賣超 3天
    Cnt3Days = 3, PatternFill(
        fill_type='solid', fgColor=colors.Color(rgb='EAFFD0'))
    # 5日內累積買或賣超 4天
    Cnt4Days = 4, PatternFill(
        fill_type='solid', fgColor=colors.Color(rgb='FCE38A'))
    # 5日內累積買或賣超 5天
    Cnt5Days = 5, PatternFill(
        fill_type='solid', fgColor=colors.Color(rgb='F38181'))


def GetColorRGB(val):
    """
    回傳顏色rgb
    val:enum value[0]
    """
    return ColorEnum(val).values[1]


class BuySell():
    """資料輸出產生excel sheet物件"""

    def __init__(self, dateList):
        """Constructor"""
        self.name = ""
        self.EnName = ""

        global sheet

        self.dList = dateList
        # 是否爬到資料
        self.IsGetSource = True
        # 爬到的初始資料
        self.DataSourceByScrapy = []
        # 從初始資料過濾取得所需欄位
        self.FilterData = []
        # 整理過濾後的預備輸出資料
        self.ExportData = []

        # 擷取指定資料筆數
        self.CutRows = 30
        # 外資買超前指定資料筆數
        self.Foreign_Buy = np.array([])
        # 外資賣超
        self.Foreign_Sell = np.array([])
        # 投信買超
        self.Local_Buy = np.array([])
        # 投信賣超
        self.Local_Sell = np.array([])

    def ScrapyData(self, dateObj):
        """爬取資料
        dataObj:日期資料物件
        """

    def FilterScrapyData(self):
        """
        資料整理
        將爬取的資料整理並指派給 FilterData
        """

    def __ArrangeData(self, target, action):
        """
        對FilterData排序並取出指定筆數的資料
        target:F外資、L投信
        action:B買超、S賣超
        """
        # 依外資或投信買賣超排序
        sortColIdx = 0
        # ture:買超排序、false:賣超排序
        sortReverse = True
        # 刪除指定欄位 ex.外資刪除投信買賣超欄位
        delColIdx = 0

        sortColIdx = target == 'F' and 2 or 3
        delColIdx = target == 'F' and 3 or 2
        sortReverse = action == 'B' and True or False

        # 依照外資或投信的買或賣超排序
        self.FilterData.sort(key=lambda x: x[sortColIdx], reverse=sortReverse)
        """
        1.
        使用copy.deepcopy 完整複製一份reference object
        因為使用list.copy() 假設新list pop時 會異動到舊list
        (只有clear()不會 append()亦會影響)

        2.
        dtype=object:為了將張數轉為數字，不然會是字串型態
        """
        npObj = np.array(copy.deepcopy(self.FilterData)
                         [:self.CutRows], dtype=object)

        """
        axis=None：arr会先按行展开，然后按照obj，删除第obj-1（从0开始）位置的数，返回一个行矩阵。
        axis = 0：arr按橫列删除
        axis = 1：arr按直行删除
        ex.外資刪除 投信買賣超欄位
        """
        npObj = np.delete(npObj, delColIdx, axis=1)

        # 刪除買賣超中 張數為0的資料列
        for idx, elem in enumerate(npObj):
            if elem[2] == 0:
                npObj = np.delete(npObj, np.s_[idx:], axis=0)
                break
        return npObj

    def SetForeignAndLocalData(self):
        """資料整理
        將爬取的資料整理並指派給 外資、投信買超和賣超 member
        取外資、投信買賣超排行前30列
        """
        # 外資買超
        self.Foreign_Buy = np.array([])
        # 外資賣超
        self.Foreign_Sell = np.array([])
        # 投信買超
        self.Local_Buy = np.array([])
        # 投信賣超
        self.Local_Sell = np.array([])
        
        self.Foreign_Buy = self.__ArrangeData('F', 'B')        
        self.Foreign_Sell = self.__ArrangeData('F', 'S')        
        self.Local_Buy = self.__ArrangeData('L', 'B')        
        self.Local_Sell = self.__ArrangeData('L', 'S')

    def SetExportData(self):
        """
        將外資、投信買賣超資料組合給ExportData
        """
        # 準備最後輸出的資料
        self.ExportData = []
        np4 = [self.Foreign_Buy, self.Foreign_Sell,
               self.Local_Buy, self.Local_Sell]
        # exportdata 遇到買賣超無筆數資料則補空白
        empty_list = ['', '', '']
        for i in range(self.CutRows):
            row_data = []
            for idx, elem in enumerate(np4):
                # 資料處理筆數index超過elem row count
                if i >= elem.shape[0]:
                    row_data.extend(empty_list)
                else:
                    row_data.extend(elem[i])
            self.ExportData.append(row_data)

    def ExportExcel(self, dataObj):
        """輸出資料至excel
        excel資料寫入
        dateObj:日期物件
        """

        # text of title's row
        resultList_tw = [
            '代號', '公司', '外資買超張數',
            '代號', '公司', '外資賣超張數',
            '代號', '公司', '投信買超張數',
            '代號', '公司', '投信賣超張數']

        fileObj = FileIO.FileIO(self.EnName, dataObj.strdate)
        row_idx = 1  # row index
        col_idx = 1  # column index

        wb = fileObj.GetWorkBook()
        sheetName = dataObj.strdate[4:]

        # 字體大小
        font12 = Font(size=12)

        # 比對同步買超、賣超使用
        # ex:[2, self.Local_Buy, self.Local_Sell]
        # 外資買超,同向比對資料內容,反向比對資料內容
        buysell = np.array([
            [2, self.Local_Buy, self.Local_Sell],
            [5, self.Local_Sell, self.Local_Buy],
            [8, self.Foreign_Buy, self.Foreign_Sell],
            [11, self.Foreign_Sell, self.Foreign_Buy],
        ])

        # 歷史資料
        historyList = HistoryFind.GetHistory(fileObj.SaveAs)

        if fileObj.XlsExist:
            if sheetName in wb.sheetnames:
                print(sheetName + self.name + dataObj.strdate + "買賣超已存在")
                # sys.exit()
                return
            else:
                sheet = wb.create_sheet(sheetName)
        else:
            # 創建一個工作表(注意是一個屬性)
            sheet = wb.active
            # excel創建的工作表名默認為sheet1,一下代碼實現了給新創建的工作表創建一個新的名字
            sheet.title = sheetName

        titleList = ['同向', '反向', '同2', '同3', '同4', '同5']
        for idx, elem in enumerate(titleList):
            sheet.cell(row=row_idx, column=col_idx).value = elem
            sheet.cell(row=row_idx, column=col_idx).fill = GetColorRGB(idx)
            col_idx += 1

        row_idx += 1
        col_idx = 1

        # 向工作表中輸入內容1-標題
        sheet.append(resultList_tw)
        row_idx += 1

        # 內容
        for row in self.ExportData:
            col_idx = 1
            for cell in row:
                sheet.cell(row=row_idx, column=col_idx).value = cell
                sheet.cell(row=row_idx, column=col_idx).font = font12
                # 買賣超張數欄位
                if (col_idx % 3) == 0:
                    quotient = int(col_idx / 3)  # 取商數
                    stockNo = sheet.cell(row=row_idx, column=(
                        col_idx - 2)).value  # 取股票代號欄位值
                    cnt = HistoryFind.GetItemCount(
                        historyList, quotient - 1, stockNo)
                    cnt += 1  # 前四日重複出現次數 + 該日顯示次數1次 = 五日內出現次數
                    sheet.cell(
                        row=row_idx, column=col_idx).number_format = '#,##0'
                    if cnt > 1:
                        # cnt 2~5 剛好等於 colorenum values[cnt]
                        sheet.cell(row=row_idx, column=col_idx).fill = GetColorRGB(
                            cnt)

                findNp = _cm.FindNumpyIdx(buysell, col_idx, 0)
                # 判定是否為買賣超張數欄位
                if np.size(findNp) > 0:
                    # 同向 (比對股票中文名稱欄位)
                    secondFindNp1 = _cm.FindNumpyIdx(findNp[0][1], cell, 1)
                    # 反向
                    secondFindNp2 = _cm.FindNumpyIdx(findNp[0][2], cell, 1)

                    if np.size(secondFindNp1) > 0:
                        sheet.cell(
                            row=row_idx, column=col_idx).fill = ColorEnum.Yello.values[1]
                    elif np.size(secondFindNp2) > 0:
                        sheet.cell(
                            row=row_idx, column=col_idx).fill = ColorEnum.Green.values[1]
                col_idx += 1
            row_idx += 1

        # 外資、投信買賣超間隔一空白欄
        sheet.insert_cols(7)

        # 欄位寬度調整
        for col in sheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                if cell.coordinate in sheet.merged_cells:  # not check merge_cells
                    continue
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 5) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        # 保存一個文檔
        wb.save(fileObj.SaveAs)

    def Execute(self):
        """
        class main methon
        """
        for d in self.dList:
            dateObj = DateObj.DateObj(d)
            # 爬取資料
            self.ScrapyData(dateObj)
            # 判斷是否有爬取到資料
            if not self.IsGetSource:
                print(self.name + "買賣超讀取失敗")
                return
                # sys.exit()
            # 資料整理
            self.FilterScrapyData()
            self.SetForeignAndLocalData()
            self.SetExportData()
            # 資料輸出
            self.ExportExcel(dateObj)
