# 上櫃外資 投信買賣超前三十
import sys
import requests
from bs4 import BeautifulSoup
import pandas as pd
import copy
from package.Infrastructure import DateObj, FileIO, HistoryFind
from openpyxl import Workbook
from openpyxl import load_workbook
# 使用openpyxl内置的格式
from openpyxl.styles import numbers, Font, PatternFill, colors

"""
使用copy.deepcopy 完整複製一份reference object
因為使用list.copy() 假設新list pop時 會異動到舊list
(只有clear()不會 append()亦會影響)
"""


def formatNo(noStr):
    noOrg = noStr.replace(',', '')
    noInt = int(noOrg)
    return int(noInt / 1000)


def FindListIdx(listObj, elem):  # list.index(ele) 假設沒找到會丟value error
    if elem.strip() == "":
        return -1

    for row, i in enumerate(listObj):
        try:
            column = i.index(elem)
        except ValueError:
            continue
        return 1
    return -1


def GetWorkBook(fileObj):
    """取得或新增 excel"""
    if fileObj.XlsExist:
        # 先備份檔案
        fileObj.CopyFile()
        return load_workbook(filename=fileObj.SaveAs)
    else:
        return Workbook()


def main(dList):
    # 日期list
    dateList = dList
    for d in dateList:
        dateObj = DateObj.DateObj(d)
        ExportExcel(dateObj)


def ExportExcel(obj):
    resultList_tw = [
        '代號', '公司', '外資買超張數',
        '代號', '公司', '外資賣超張數',
        '代號', '公司', '投信買超張數',
        '代號', '公司', '投信賣超張數']

    url = 'http://www.tpex.org.tw/web/stock/3insti/daily_trade/3itrade_hedge_result.php'
    data = {
        'l': 'zh-tw',
        'se': 'EW',  # AL:全部、EW:不含權證、牛熊證
        't': 'D',
        'd': obj.twSlashDate
    }
    r = requests.get(url, params=data)
    data_json = r.json()
    basic_data = data_json["aaData"]
    adjust_data = []
    if len(basic_data) == 0:
        print("上櫃買賣超讀取失敗")
        sys.exit()

    for x in basic_data:
        one_data = [str(x[0]).strip(), x[1].strip(),
                    formatNo(x[10]),  # 外陸資買賣超股數
                    formatNo(x[13])]  # 投信買賣超股數
        adjust_data.append(one_data)

    foreign_Buy = []  # 外資買超
    foreign_Sell = []  # 外資賣超
    local_Buy = []  # 投信買超
    local_Sell = []  # 投信賣超

    # 外資買超前三十
    adjust_data.sort(key=lambda x: x[2], reverse=True)
    foreign_Buy = copy.deepcopy(adjust_data)[:30]

    # 外資賣超
    adjust_data.sort(key=lambda x: x[2], reverse=False)
    foreign_Sell = copy.deepcopy(adjust_data)[:30]

    # 投信買超
    adjust_data.sort(key=lambda x: x[3], reverse=True)
    local_Buy = copy.deepcopy(adjust_data)[:30]

    # 投信賣超
    adjust_data.sort(key=lambda x: x[3], reverse=False)
    local_Sell = copy.deepcopy(adjust_data)[:30]

    final_data = []
    empty_list = ['', '', '']
    for x in foreign_Buy:
        idx = foreign_Buy.index(x)
        row_data = []
        # 外資買超
        foreign_Buy[idx].pop(3)
        if foreign_Buy[idx][2] == 0:
            row_data.extend(empty_list)
        else:
            row_data.extend(foreign_Buy[idx])

        # 外資賣超
        foreign_Sell[idx].pop(3)
        if foreign_Sell[idx][2] == 0:
            row_data.extend(empty_list)
        else:
            row_data.extend(foreign_Sell[idx])

        # 投信買超
        local_Buy[idx].pop(2)
        if local_Buy[idx][2] == 0:
            row_data.extend(empty_list)
        else:
            row_data.extend(local_Buy[idx])

        # 投信賣超
        local_Sell[idx].pop(2)
        if local_Sell[idx][2] == 0:
            row_data.extend(empty_list)
        else:
            row_data.extend(local_Sell[idx])

        final_data.append(row_data)

    # 去掉買賣超中 張數為0的資料列
    for l in [foreign_Buy, foreign_Sell, local_Buy, local_Sell]:
        for x in l:
            idx = l.index(x)
            if l[idx][2] == 0:
                del l[idx:]
                break

# openpyxl export
    # 檔案操作
    fileObj = FileIO.FileIO('OTCBuySell', obj.strdate)
    row_idx = 1
    col_idx = 1
    yellofill = PatternFill(fill_type='solid', fgColor=colors.YELLOW)
    greenfill = PatternFill(fill_type='solid', fgColor=colors.GREEN)
    day2Color = colors.Color(rgb='0099FFFF')
    day2fill = PatternFill(fill_type='solid', fgColor=day2Color)
    day3Color = colors.Color(rgb='0099CCFF')
    day3fill = PatternFill(fill_type='solid', fgColor=day3Color)
    day4Color = colors.Color(rgb='009999FF')
    day4fill = PatternFill(fill_type='solid', fgColor=day4Color)
    day5Color = colors.Color(rgb='00CC99FF')
    day5fill = PatternFill(fill_type='solid', fgColor=day5Color)

    wb = GetWorkBook(fileObj)
    sheetName = obj.strdate[4:]
    global sheet
    # 歷史資料
    historyList = HistoryFind.GetHistory(fileObj.SaveAs)

    if fileObj.XlsExist:
        if sheetName in wb.sheetnames:
            print(sheetName + "上櫃買賣超已存在")
            sys.exit()
        else:
            sheet = wb.create_sheet(sheetName)
    else:
        # 創建一個工作表(注意是一個屬性)
        sheet = wb.active
        # excel創建的工作表名默認為sheet1,一下代碼實現了給新創建的工作表創建一個新的名字
        sheet.title = sheetName

    sheet.cell(row=row_idx, column=col_idx).value = "同向"
    sheet.cell(row=row_idx, column=col_idx).fill = yellofill

    col_idx += 1
    sheet.cell(row=row_idx, column=col_idx).value = "反向"
    sheet.cell(row=row_idx, column=col_idx).fill = greenfill

    col_idx += 1
    sheet.cell(row=row_idx, column=col_idx).value = "同2"
    sheet.cell(row=row_idx, column=col_idx).fill = day2fill

    col_idx += 1
    sheet.cell(row=row_idx, column=col_idx).value = "同3"
    sheet.cell(row=row_idx, column=col_idx).fill = day3fill

    col_idx += 1
    sheet.cell(row=row_idx, column=col_idx).value = "同4"
    sheet.cell(row=row_idx, column=col_idx).fill = day4fill

    col_idx += 1
    sheet.cell(row=row_idx, column=col_idx).value = "同5"
    sheet.cell(row=row_idx, column=col_idx).fill = day5fill

    row_idx += 1
    col_idx = 1

    # 向工作表中輸入內容1-標題
    sheet.append(resultList_tw)

    row_idx += 1
    # 內容
    for row in final_data:
        col_idx = 1
        for cell in row:
            sheet.cell(row=row_idx, column=col_idx).value = cell
            if (col_idx % 3) == 0:
                quotient = int(col_idx / 3)
                stockNo = sheet.cell(row=row_idx, column=(
                    (quotient - 1)*3) + 1).value
                cnt = HistoryFind.GetItemCount(
                    historyList, quotient - 1, stockNo)
                cnt += 1
                sheet.cell(
                    row=row_idx, column=col_idx).number_format = '#,##0'
                if cnt == 2:
                    sheet.cell(row=row_idx, column=col_idx).fill = day2fill
                if cnt == 3:
                    sheet.cell(row=row_idx, column=col_idx).fill = day3fill
                if cnt == 4:
                    sheet.cell(row=row_idx, column=col_idx).fill = day4fill
                if cnt == 5:
                    sheet.cell(row=row_idx, column=col_idx).fill = day5fill

            if col_idx == 2:  # 外資買超
                if FindListIdx(local_Buy, cell) > 0:
                    sheet.cell(
                        row=row_idx, column=col_idx).fill = yellofill
                elif FindListIdx(local_Sell, cell) > 0:
                    sheet.cell(
                        row=row_idx, column=col_idx).fill = greenfill
            if col_idx == 5:  # 外資賣超
                if FindListIdx(local_Sell, cell) > 0:
                    sheet.cell(
                        row=row_idx, column=col_idx).fill = yellofill
                elif FindListIdx(local_Buy, cell) > 0:
                    sheet.cell(
                        row=row_idx, column=col_idx).fill = greenfill

            if col_idx == 8:  # 投信買超
                if FindListIdx(foreign_Buy, cell) > 0:
                    sheet.cell(
                        row=row_idx, column=col_idx).fill = yellofill
                elif FindListIdx(foreign_Sell, cell) > 0:
                    sheet.cell(
                        row=row_idx, column=col_idx).fill = greenfill
            if col_idx == 11:  # 投信賣超
                if FindListIdx(foreign_Sell, cell) > 0:
                    sheet.cell(
                        row=row_idx, column=col_idx).fill = yellofill
                elif FindListIdx(foreign_Buy, cell) > 0:
                    sheet.cell(
                        row=row_idx, column=col_idx).fill = greenfill
            col_idx += 1
        row_idx += 1
    # 保存一個文檔
    wb.save(fileObj.SaveAs)
