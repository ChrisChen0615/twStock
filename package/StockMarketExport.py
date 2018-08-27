# 主程式
from package.Infrastructure import DateObj, FileIO
from package.Infrastructure import CommonMethon as _cm
from package.ScrapySource import Taeix, ForeignInvestor, PcRatio, MTX
from package.ScrapySource import ForeignFuture, StockOption, FiveAndTen
from openpyxl import Workbook
# 使用openpyxl内置的格式
from openpyxl.styles import numbers, Font, Alignment


def ScrapyData(dateList):
    """
    取得excel各欄位所需資料
    dateList:日期list
    """
    # 最後輸出結果
    resultList = []

    for d in dateList:
        dateObj = DateObj.DateObj(d)
        singleDayList = []
        singleDayList.append(dateObj.dateSlash)

        # '加權指數', '漲跌指數', '漲跌趴數', '成交量'
        t = Taeix.Taeix(dateObj)
        tList = t.GetTaeixList()
        for idx in range(0, len(tList[:-1])):
            tList[idx] = tList[idx].replace(',', '')
            tList[idx] = _cm.formatNumType(
                tList[idx], "float")  # string to float
        tList[-1] = _cm.formatNumType(tList[-1], "decimal")  # float to decimal
        singleDayList.extend(tList)

        # '外資買賣超'
        fi = ForeignInvestor.ForeignInvestor(dateObj)
        fiStr = _cm.formatNumType(fi.GetCount(), "decimal")
        singleDayList.append(fiStr)

        # 'P/C ratio'
        p = PcRatio.PCRatio(dateObj)
        pStr = p.GetRatio()
        singleDayList.append(pStr)

        # '小台指未沖銷契約量, 散戶看多, 散戶看空, 散戶多空比'
        m = MTX.MTX(dateObj)
        m.CalCount()
        singleDayList.append(m.ratio)

        ff = ForeignFuture.ForeignFuture(dateObj)
        ff.CalCount()
        singleDayList.append(_cm.formatNumType(ff.future, "int"))

        # '自營(選)買權', '自營(選)賣權', '外資(選)買權', '外資(選)賣權'
        s = StockOption.StockOption(dateObj)
        s.CalCount()
        singleDayList.append(_cm.formatNumType(s.dealer_call, "int"))
        singleDayList.append(_cm.formatNumType(s.dealer_put, "int"))
        singleDayList.append(_cm.formatNumType(s.foreign_call, "int"))
        singleDayList.append(_cm.formatNumType(s.foreign_put, "int"))

        # '前五大交易人留倉部位(所有)', '前10大交易人留倉部位(所有)'
        fat = FiveAndTen.FiveAndTen(dateObj)
        fat.CalCount()
        singleDayList.append(_cm.formatNumType(fat.five_all, "int"))
        singleDayList.append(_cm.formatNumType(fat.ten_all, "int"))

        resultList.append(singleDayList)

    return resultList


def ExportExcel(dateList, resultList):
    """
    資料輸出至excel
    dateList:日期list
    resultList:爬取後的資料
    """
    # 最後輸出結果標題列
    resultList_tw = [
        '日期',
        '加權指數', '漲跌指數', '漲跌趴數', '成交量(億)',
        '外資買賣超(億)', 'P/C ratio',
        # '小台指未沖銷契約量', '散戶看多', '散戶看空', '散戶多空比',
        '散戶多空比\n(口數)',
        '外資未平倉口數\n(大小台合計)',
        '自營(選)\n買權', '自營(選)\n賣權', '外資(選)\n買權', '外資(選)\n賣權',
        '前五大交易人\n留倉部位(所有)', '前10大交易人\n留倉部位(所有)']

    # 檔案操作
    fileObj = FileIO.FileIO('TWSE', dateList[0])

    row_idx = 1  # 列 起始:1
    col_idx = 1  # 行 起始:1

    # cell style
    alignment = Alignment(wrap_text=True, vertical='center')
    font = Font(color="FF0000")  # cell font color

    wb = fileObj.GetWorkBook()
    global sheet
    if fileObj.XlsExist:
        sheet = wb['大盤']
    else:
        # 創建一個工作表(注意是一個屬性)
        sheet = wb.active
        # excel創建的工作表名默認為sheet1,一下代碼實現了給新創建的工作表創建一個新的名字
        sheet.title = '大盤'
        # 向工作表中輸入內容1-標題(自動換行)
        sheet.append(resultList_tw)

        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                # cell.style.alignment.wrap_text = True #no alignment attr
                cell.alignment = alignment

    row_idx += 1

    # 內容
    # cell value type皆是float,decimal，再formatting成指定格式
    for row in resultList:
        sheet.insert_rows(2)
        col_idx = 1
        for r in row:
            # 離最近日期開始排序，所以從標題下一列開始寫入資料
            sheet.cell(row=2, column=col_idx).value = r
            if col_idx > 1:
                if (col_idx == 4) or (col_idx == 7) or (col_idx == 8):
                    sheet.cell(
                        row=2, column=col_idx).number_format = numbers.FORMAT_PERCENTAGE_00
                    if col_idx == 7:  # pc/ratio 小於100%偏空 紅字顯示
                        if r < 1:
                            sheet.cell(row=2, column=col_idx).font = font

                elif col_idx >= 9:
                    sheet.cell(
                        row=2, column=col_idx).number_format = '#,##0'
                else:
                    sheet.cell(
                        row=2, column=col_idx).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1  # '#,##0.00'
                if r < 0:  # 小於0偏空 紅字顯示
                    sheet.cell(row=2, column=col_idx).font = font
            col_idx += 1

    # 保存一個文檔
    wb.save(fileObj.SaveAs)


def main(dList):
    # 日期list
    dateList = dList
    # 最後輸出結果
    resultList = ScrapyData(dateList)
    # 輸出至excel
    ExportExcel(dateList, resultList)