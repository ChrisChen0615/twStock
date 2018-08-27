# -*- coding: utf-8 -*-
# 取得歷史買賣超紀錄、計算前四日出現次數
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from package.Infrastructure import FileIO


def GetHistory(filePath):
    """
    取得前四天外資、投信買賣超紀錄
    filePath:檔案路徑
    """
    if not os.path.isfile(filePath):  # 檔案是否存在
        return []

    wb = load_workbook(filename=filePath)
    wbList = wb.sheetnames
    # 最多從後面取4個sheet
    overFive = len(wbList) >= 5 and wbList[-4:] or []
    sheet4List = []  # 買賣超股票代號list(最多4個sheet)
    for sheetName in overFive:
        one_sheet = wb[sheetName]
        # 買賣超股票代號list(單sheet)
        sheetList = []

        """
        取得各買賣超股票代號list
        colIdx:外資買超、外資賣超、投信買超、投信賣超股票代號欄位數
        """
        for colIdx in [1, 4, 8, 11]:
            # 買賣超股票代號list(單欄)
            data = [one_sheet.cell(row=i, column=colIdx).value for i in range(3, 33) if str(one_sheet.cell(
                row=i, column=colIdx).value).strip() != "None"]
            sheetList.append(data)

        sheet4List.append(sheetList)

    return sheet4List


def GetItemCount(listObj, typeObj, val):
    """
    搜尋出現次數
    listObj:sheet list
    typeObj:外資買超0、外資賣超1、投信買超2、投信賣超3
    val:股票代號
    """
    cnt = 0
    for l in listObj:
        cnt += l[typeObj].count(val)
    return cnt
